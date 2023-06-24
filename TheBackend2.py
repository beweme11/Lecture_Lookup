
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
"""workbook = openpyxl.load_workbook('time_table.xlsx')

for sheet in workbook:
    sheet.sheet_state = 'visible'

for sheet in workbook: #loop
    if sheet['A1'].value != "DAY": #if not a1 value = DAY them
        workbook.remove(sheet) #remove/delete

# Save the updated workbook
workbook.save('timetable_updated.xlsx') #save the updated workbook"""
workbook2 = openpyxl.load_workbook('timetable_updated.xlsx')

"""
table_list = ('Table 2',
'Table 7',
'Table 11',
'Table 15',
'Table 20',
'Table 25',
'Table 30',
'Table 35',
'Table 40',
'Table 45',
'Table 50',
'Table 54',
'Table 58',
'Table 63',
'Table 68',
'Table 73')"""

"""print(table_list[div_index])"""

"""for sheet_name in workbook2.sheetnames : #to print the sheetnames
    print(sheet_name)"""

"""sheet_A = workbook2.worksheets[0] #accessed the 1st worksheet using index
cell = sheet_A['C3']
print(cell.value)"""
"""
time_no_dict = {'8:00': '1',
                '9:00': '2',
                '10:15': '3',
                '11:15': '4',
                '1:15' : '5',
                '2:15' : '6',
                '3:30': '7',
                '4:30' : '8',
                '5:30' : '9'}
no_column_dict = {
    '1': 'B',
    '2': 'D',
    '3': 'G',
    '4': 'I',
    '5': 'L',
    '6': 'N',
    '7': 'Q',
    '8': 'S',
    '9': 'U'
}"""

time_list = ['8:00',
             '9:00',
             '10:15',
             '11:15',
             '1:15',
             '2:15',
             '3:30',
             '4:30',
             '5:30']

column_list = [
    'B',
    'D',
    'G',
    'I',
    'L',
    'N',
    'Q',
    'S',
    'U']


row_dict = {
    'MON': '3',
    'TUE': '9',
    'WED': '15',
    'THU': '21',
    'FRI': '27',
    'SAT': '33',
    'SUN': '39'
}
row_dict2 = {
    'MON': '4',
    'TUE': '10',
    'WED': '16',
    'THU': '22',
    'FRI': '28',
    'SAT': '34',
    'SUN': '40'
}
batches = ['A1',
           'A2',
           'A3',
           'B1',
           'B2', 'B3', 'C1', 'C2', 'C3', 'D1', 'D2', 'D3', 'E1', 'E2', 'E3', 
           'F1', 'F2', 'F3', 'G1', 'G2', 'G3', 'H1', 'H2', 'H3', 'I1', 'I2', 'I3',
           'J1', 'J2', 'J3', 'K1', 'K2', 'K3', 'L1', 'L2', 'L3', 'M1', 'M2', 'M3', 
           'N1', 'N2', 'N3', 'O1', 'O2', 'O3', 'P1', 'P2', 'P3', 'Q1', 'Q2', 'Q3', 
           ]

def timetable_teller(div, batch):
    if batch in batches:
        day = input("What day are you looking for? Please enter the first 3 letters of the day only (example: MON): ")
        time = input("What time class are you looking for? Please enter in 12hr format (example: 10:30): ")
        #To get the division
        div_list = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P')
        div = input("Enter division: ")
        batch = input("Enter batch: ")
        sheet_index = div_list.index(div)
        sheet_A = workbook2.worksheets[sheet_index]
        
        # For lecture timetable
        column_index_lecture = column_index_from_string(column_list[time_list.index(time)])
        next_column_lecture = get_column_letter(column_index_lecture + 1)
        cell_of_interest_lecture = next_column_lecture + row_dict[day]
        value_of_cell_of_interest_lecture = sheet_A[cell_of_interest_lecture].value
        print("The lecture is :",value_of_cell_of_interest_lecture)
        
        # For class timetable
        column_index = column_index_from_string(column_list[time_list.index(time)])
        next_column = get_column_letter(column_index)
        cell_of_interest = next_column + row_dict2[day]
        value_of_cell_of_interest_class = sheet_A[cell_of_interest].value
        print("The Class room is ", value_of_cell_of_interest_class)
        class_status = {
            "E305": "Projector screen broken",
            "E306": "Air conditioning not working",
            "E312": "Classroom rearranged"
        }
        
        if value_of_cell_of_interest_class in class_status:
            print(f"The Status/Condition of {value_of_cell_of_interest_class} Class is:" , class_status[value_of_cell_of_interest_class])
        else:
           print(f"The Status/Condition of {value_of_cell_of_interest_class} Class is Good")
    else:
        print("Invalid batch value.")
    
timetable_teller()

